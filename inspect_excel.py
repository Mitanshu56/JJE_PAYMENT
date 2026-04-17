from openpyxl import load_workbook

wb = load_workbook('Copy of bill 24-25(1).xlsx', data_only=True)
print('Sheets:', wb.sheetnames)

for s in wb.sheetnames:
    ws = wb[s]
    print(f"\n=== {s} ===")
    for r in range(1, 26):
        vals = [ws.cell(r, c).value for c in range(1, 13)]
        if any(v is not None and str(v).strip() != '' for v in vals):
            print(r, vals)
