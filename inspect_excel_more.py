from openpyxl import load_workbook

wb = load_workbook('Copy of bill 24-25(1).xlsx', data_only=True)
for s in ['24-25','havells bill']:
    ws = wb[s]
    print(f"\n=== {s} rows 19-80 ===")
    for r in range(19, 81):
        vals = [ws.cell(r, c).value for c in range(1, 13)]
        if any(v is not None and str(v).strip() != '' for v in vals):
            print(r, vals)
