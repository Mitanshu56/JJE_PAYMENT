from openpyxl import load_workbook

wb = load_workbook(r'Copy of bill 25-26.xlsx', data_only=True)
print('sheets', wb.sheetnames)
first = wb[wb.sheetnames[0]]
print('first_sheet', wb.sheetnames[0])
for r in range(1, 25):
    vals = [first.cell(r, c).value for c in range(1, 10)]
    if any(v is not None and str(v).strip() for v in vals):
        print(r, vals)
